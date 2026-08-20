"""Create the synthetic Excel workbook used in the PromptJeopardy workshop."""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


WORKSPACE = Path(__file__).resolve().parents[1]
OUTPUT = WORKSPACE / "public" / "ovningsfiler" / "Semesterplan-demo.xlsx"

NAVY = "0F1B3D"
BLUE = "2563EB"
LIGHT_BLUE = "DBEAFE"
GREEN = "047857"
LIGHT_GREEN = "D1FAE5"
YELLOW = "FACC15"
LIGHT_YELLOW = "FEF3C7"
RED = "B91C1C"
LIGHT_RED = "FEE2E2"
SLATE = "334155"
LIGHT_SLATE = "E2E8F0"
WHITE = "FFFFFF"
THIN_GREY = Side(style="thin", color="CBD5E1")

EMPLOYEES = [
    ("Alva Berg", "Service", "Handläggare", 100),
    ("Samir Demir", "Service", "Handläggare", 100),
    ("Linnea Falk", "Service", "Specialist", 100),
    ("Amir Gashi", "Service", "Handläggare", 100),
    ("Elena Holm", "Service", "Teamledare", 100),
    ("Johan Isaksson", "Analys", "Analytiker", 100),
    ("Miriam Jallow", "Analys", "Analytiker", 100),
    ("Leo Karlsson", "Analys", "Specialist", 100),
    ("Amina Lind", "Analys", "Analytiker", 100),
    ("Sofia Moberg", "Analys", "Teamledare", 100),
    ("Viktor Norén", "Digitalt", "Utvecklare", 100),
    ("Nora Osman", "Digitalt", "Utvecklare", 100),
    ("Omar Persson", "Digitalt", "Produktägare", 100),
    ("Elin Qvist", "Digitalt", "Utvecklare", 100),
    ("David Rahm", "Digitalt", "Specialist", 100),
    ("Fatima Said", "Stöd", "Kommunikatör", 100),
    ("Magnus Tran", "Stöd", "Controller", 100),
    ("Sara Vik", "Stöd", "HR-specialist", 100),
    ("Rami Wallin", "Stöd", "Jurist", 100),
    ("Hanna Öberg", "Stöd", "Teamledare", 100),
]

VACATIONS = {
    0: (date(2026, 7, 13), date(2026, 7, 24)),
    1: (date(2026, 7, 13), date(2026, 7, 24)),
    2: (date(2026, 7, 13), date(2026, 7, 24)),
    3: (date(2026, 7, 20), date(2026, 7, 31)),
    4: (date(2026, 7, 20), date(2026, 7, 31)),
    5: (date(2026, 7, 20), date(2026, 7, 31)),
    6: (date(2026, 7, 20), date(2026, 7, 31)),
    7: (date(2026, 7, 20), date(2026, 7, 24)),
    8: (date(2026, 7, 20), date(2026, 7, 24)),
    9: (date(2026, 7, 22), date(2026, 7, 24)),
    10: (date(2026, 7, 22), date(2026, 7, 24)),
    11: (date(2026, 7, 22), date(2026, 7, 24)),
    12: (date(2026, 7, 22), date(2026, 7, 23)),
    13: (date(2026, 7, 27), date(2026, 8, 7)),
    14: (date(2026, 7, 27), date(2026, 8, 7)),
    15: (date(2026, 6, 29), date(2026, 7, 10)),
    16: (date(2026, 6, 29), date(2026, 7, 10)),
    17: (date(2026, 7, 6), date(2026, 7, 17)),
    18: (date(2026, 8, 3), date(2026, 8, 7)),
    19: (date(2026, 8, 3), date(2026, 8, 7)),
}

SPECIAL_STATUSES = {
    (19, date(2026, 7, 20)): "Sjuk",
    (19, date(2026, 7, 22)): "Sjuk",
    (9, date(2026, 7, 21)): "Utbildning",
    (10, date(2026, 7, 21)): "Halvdag",
}


def working_days(start: date, end: date) -> list[date]:
    days: list[date] = []
    current = start
    while current <= end:
        if current.weekday() < 5:
            days.append(current)
        current += timedelta(days=1)
    return days


def employee_status(employee_index: int, day: date) -> str:
    special = SPECIAL_STATUSES.get((employee_index, day))
    if special:
        return special
    vacation_start, vacation_end = VACATIONS[employee_index]
    if vacation_start <= day <= vacation_end:
        return "Semester"
    return "På plats"


def true_staffing(day: date) -> float:
    weights = {"På plats": 1, "Halvdag": 0.5}
    return sum(weights.get(employee_status(index, day), 0) for index in range(len(EMPLOYEES)))


def style_title(cell, size: int = 22) -> None:
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=size)
    cell.alignment = Alignment(vertical="center")


def style_header_row(sheet, row: int, start_column: int, end_column: int) -> None:
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(row=row, column=column)
        cell.fill = PatternFill("solid", fgColor=SLATE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GREY)


def add_read_me(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Läs mig"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 4
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 76
    sheet.column_dimensions["D"].width = 4

    sheet.merge_cells("B2:C2")
    sheet["B2"] = "Semesterplan – övningsfil för Prompt-Jeopardy"
    style_title(sheet["B2"])
    sheet.row_dimensions[2].height = 38

    sheet.merge_cells("B4:C4")
    sheet["B4"] = "Fiktiva uppgifter – inga personuppgifter eller riktiga verksamhetsdata"
    sheet["B4"].fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
    sheet["B4"].font = Font(color=NAVY, bold=True, size=12)
    sheet["B4"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[4].height = 32

    sections = [
        (
            6,
            "Så använder ni filen",
            "Spara en kopia i OneDrive eller SharePoint och öppna den i Excel. "
            "Använd Copilot i Excel för att undersöka arbetsboken. Ändra inget innan ni har granskat förslaget manuellt.",
        ),
        (
            9,
            "Övning 1: Formeldetektiven",
            "Leta efter formler som avviker från mönstret, felvärden och hårdkodade tal där det borde finnas en formel. "
            "Välj två misstänkta celler och förklara hur ni kontrollerade dem.",
        ),
        (
            12,
            "Övning 2: Bemanningsanalys",
            "Hitta de tre arbetsdagar som har lägst bemanning, markera dagar under minimibemanningen och skapa en enkel visualisering. "
            "Beskriv också minst en sak som underlaget inte tar hänsyn till.",
        ),
        (
            15,
            "Statuskoder",
            "På plats = 1 person  •  Halvdag = 0,5 person  •  Semester, Sjuk och Utbildning = 0 personer på plats.",
        ),
        (
            18,
            "Viktigt",
            "Det finns avsiktliga avvikelser i arbetsboken. De är en del av övningen. Be Copilot förklara sina fynd och kontrollera alltid celler, formler och källdata själv.",
        ),
    ]
    for row, heading, body in sections:
        sheet[f"B{row}"] = heading
        sheet[f"B{row}"].font = Font(color=BLUE, bold=True, size=13)
        sheet[f"C{row}"] = body
        sheet[f"C{row}"].font = Font(color=NAVY, size=11)
        sheet[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 45 if row != 15 else 32

    sheet["B21"] = "Minimibemanning"
    sheet["C21"] = "10 personer på plats per arbetsdag"
    sheet["B21"].font = Font(color=GREEN, bold=True, size=13)
    sheet["C21"].font = Font(color=NAVY, bold=True, size=12)


def add_schedule(workbook: Workbook, days: list[date]) -> None:
    sheet = workbook.create_sheet("Semesterplan")
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "E5"

    end_column = 4 + len(days)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    sheet.cell(row=1, column=1, value="SEMESTERPLAN – FIKTIVT ÖVNINGSMATERIAL")
    style_title(sheet.cell(row=1, column=1))
    sheet.row_dimensions[1].height = 38
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    sheet.cell(
        row=2,
        column=1,
        value="Status per arbetsdag. På plats = 1, Halvdag = 0,5. Övriga statusar räknas inte som bemanning.",
    )
    sheet.cell(row=2, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet.cell(row=2, column=1).font = Font(color=NAVY, italic=True)

    headers = ["Namn", "Team", "Roll", "Tjänstgöring"] + [day.isoformat() for day in days]
    for column, header in enumerate(headers, 1):
        sheet.cell(row=4, column=column, value=header)
    style_header_row(sheet, 4, 1, end_column)
    sheet.row_dimensions[4].height = 48

    for employee_index, (name, team, role, employment) in enumerate(EMPLOYEES):
        row = 5 + employee_index
        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=2, value=team)
        sheet.cell(row=row, column=3, value=role)
        sheet.cell(row=row, column=4, value=employment / 100)
        sheet.cell(row=row, column=4).number_format = "0%"
        for day_index, day in enumerate(days):
            column = 5 + day_index
            cell = sheet.cell(row=row, column=column, value=employee_status(employee_index, day))
            cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)

    table = Table(displayName="SemesterplanTabell", ref=f"A4:{get_column_letter(end_column)}24")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    sheet.add_table(table)

    validation = DataValidation(
        type="list",
        formula1='"På plats,Halvdag,Semester,Sjuk,Utbildning"',
        allow_blank=False,
    )
    validation.error = "Välj en status i listan."
    validation.errorTitle = "Ogiltig status"
    sheet.add_data_validation(validation)
    validation.add(f"E5:{get_column_letter(end_column)}24")

    fills = {
        "På plats": LIGHT_GREEN,
        "Halvdag": LIGHT_YELLOW,
        "Semester": LIGHT_BLUE,
        "Sjuk": LIGHT_RED,
        "Utbildning": LIGHT_SLATE,
    }
    for status, color in fills.items():
        sheet.conditional_formatting.add(
            f"E5:{get_column_letter(end_column)}24",
            FormulaRule(formula=[f'E5="{status}"'], fill=PatternFill("solid", fgColor=color)),
        )

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 17
    sheet.column_dimensions["D"].width = 14
    for column in range(5, end_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 6
    for row in range(5, 25):
        sheet.row_dimensions[row].height = 72


def add_staffing_overview(workbook: Workbook, days: list[date]) -> None:
    sheet = workbook.create_sheet("Bemanningsöversikt")
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "BEMANNINGSÖVERSIKT"
    style_title(sheet["A1"])
    sheet.row_dimensions[1].height = 38
    sheet.merge_cells("A2:F2")
    sheet["A2"] = "Minimibemanning: 10 personer. Formlerna hämtar status från bladet Semesterplan."
    sheet["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet["A2"].font = Font(color=NAVY, italic=True)

    headers = ["Datum", "Veckodag", "På plats", "Minimikrav", "Differens", "Bedömning"]
    for column, header in enumerate(headers, 1):
        sheet.cell(row=4, column=column, value=header)
    style_header_row(sheet, 4, 1, 6)

    weekdays = ["måndag", "tisdag", "onsdag", "torsdag", "fredag"]
    for index, day in enumerate(days):
        row = 5 + index
        schedule_column = get_column_letter(5 + index)
        cell_range = f"'Semesterplan'!{schedule_column}$5:{schedule_column}$24"
        sheet.cell(row=row, column=1, value=day)
        sheet.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        sheet.cell(row=row, column=2, value=weekdays[day.weekday()])
        sheet.cell(
            row=row,
            column=3,
            value=f'=COUNTIF({cell_range},"På plats")+0.5*COUNTIF({cell_range},"Halvdag")',
        )
        sheet.cell(row=row, column=4, value=10)
        sheet.cell(row=row, column=5, value=f"=C{row}-D{row}")
        sheet.cell(row=row, column=6, value=f'=IF(E{row}<0,"UNDER MINIMUM","OK")')

    # Three deliberate anomalies for the formula detective exercise.
    sheet["C11"] = '=COUNTIF(\'Semesterplan\'!K$5:K$23,"På plats")+0.5*COUNTIF(\'Semesterplan\'!K$5:K$23,"Halvdag")'
    sheet["C18"] = true_staffing(days[13])
    sheet["F20"] = '=IF(E20<=0,"UNDER MINIMUM","OK")'

    table = Table(displayName="BemanningsoversiktTabell", ref=f"A4:F{4 + len(days)}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    sheet.add_table(table)

    end_row = 4 + len(days)
    sheet.conditional_formatting.add(
        f"C5:C{end_row}",
        CellIsRule(operator="lessThan", formula=["10"], fill=PatternFill("solid", fgColor=LIGHT_RED), font=Font(color=RED, bold=True)),
    )
    sheet.conditional_formatting.add(
        f"F5:F{end_row}",
        FormulaRule(formula=['F5="UNDER MINIMUM"'], fill=PatternFill("solid", fgColor=LIGHT_RED), font=Font(color=RED, bold=True)),
    )

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Bemanning per arbetsdag"
    chart.y_axis.title = "Personer på plats"
    chart.x_axis.title = "Datum"
    data = Reference(sheet, min_col=3, min_row=4, max_row=end_row)
    categories = Reference(sheet, min_col=1, min_row=5, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 10
    chart.width = 22
    sheet.add_chart(chart, "H4")

    widths = {"A": 14, "B": 14, "C": 13, "D": 13, "E": 13, "F": 20}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in range(5, end_row + 1):
        for column in range(1, 7):
            sheet.cell(row=row, column=column).alignment = Alignment(vertical="center")


def add_metadata(workbook: Workbook) -> None:
    properties = workbook.properties
    properties.title = "Semesterplan-demo – PromptJeopardy"
    properties.subject = "Syntetisk Excelövning för formelgranskning och bemanningsanalys"
    properties.creator = "PromptJeopardy"
    properties.description = "Fiktiva uppgifter. Inga personuppgifter eller verkliga verksamhetsdata."
    properties.keywords = "PromptJeopardy, Copilot, Excel, workshop, syntetiska data"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"


def validate_workbook(path: Path, days: list[date]) -> None:
    workbook = load_workbook(path, data_only=False)
    assert workbook.sheetnames == ["Läs mig", "Semesterplan", "Bemanningsöversikt"]
    overview = workbook["Bemanningsöversikt"]
    assert overview["C11"].data_type == "f"
    assert overview["C18"].data_type == "n"
    assert overview["F20"].data_type == "f"
    staffing = sorted(((true_staffing(day), day) for day in days), key=lambda item: (item[0], item[1]))
    expected = [
        (6, date(2026, 7, 22)),
        (7, date(2026, 7, 23)),
        (8, date(2026, 7, 24)),
    ]
    assert staffing[:3] == expected, staffing[:3]


def main() -> None:
    days = working_days(date(2026, 6, 29), date(2026, 8, 7))
    workbook = Workbook()
    add_read_me(workbook)
    add_schedule(workbook, days)
    add_staffing_overview(workbook, days)
    add_metadata(workbook)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    validate_workbook(OUTPUT, days)
    print(f"Created {OUTPUT}")
    print("Lowest staffing: 2026-07-22 = 6, 2026-07-23 = 7, 2026-07-24 = 8")


if __name__ == "__main__":
    main()
